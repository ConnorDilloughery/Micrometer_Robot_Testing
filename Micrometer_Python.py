

# Author: Connor Dilloughery
# Company: SPT Microtechnologies USA, Inc.
# Purpose: This code will only work for MetraLight's Rx Laser Micrometers. The purpose for this code is to take measurement values as a way to see if the robot cycles can perform accurate and precise movements. 
#          This code requires a desired number of cycles and store the same amount of data points. When the desired number of cycles has been completed, 
#          the data information is stored on excel.  


import serial                                      # Used to communicate with the sensors
import time
import openpyxl


def main():
    state = 0
    init_count = 0
    reading1_list1 = []                              # List of values used to store measurement values for sensor 1
    reading2_list1 = []                              # List of values used to store measurement values for sensor 2
    reading1_list2 = []
    reading2_list2 = []
    raised_flag1 = 0                                # Flag used to see if measurement value is ready to be taken
    raised_flag2 = 0


    while True:
        if state == 0:
        # State 0: Initialization

            state = 1

        elif state == 1: 
        # State 1: Take Measurement 
        # We are opening the serial ports and taking 16 measurement values by sending 'b\x14'. 
        # The 4 in that message tells the serial ports that 2^4 measurements should be taken. 
        # These values are averaged so that we obtain a more accurate representation of what the actual measurement is.

            if init_count == 0:                                             # Checks to see if the serial ports needs to be initialized
                try:
                    ser1 = serial.Serial('COM4', 115200, timeout=1)         # Opens serial 1, Baudrate = 115200, using COM port 4 (May need to change if using other computer)
                    ser2 = serial.Serial('COM5', 115200, timeout=1)         # Opens serial 2, Baudrate = 115200, using COM port 5 (May need to change if using other computer)

                    ser1.write(b'\x34')                                     # Sets the mode for sensor 1 by assigning it the CENTER mode
                    serial_input1 = ser1.read(1)                            # Used to check that the sensor is reading the correct mode

                    ser2.write(b'\x34')                                     # Sets the mode for sensor 2 by assigning it the CENTER mode
                    serial_input2 = ser2.read(1)                            # Used to check that the sensor is reading the correct mode
                    init_count +=1                                          # Use to raise flag that we have initialized the sensors
                    state = 1
                except: 
                    ser1.close()                                            # Closes serial port 1
                    ser2.close()                                            # Closes serial port 2
            

            # Serial 1
            ser1.write(b'\x14')                                             # The 4 represents 2^4 captures from the micrometer sensor
            n = 0                                                           # Indexing variable for sensor 1
            reading1_avg = 0                                                # Variable used to obtain our measurement value
        
            while n < 16:                                                   # 16 readings since 2^4, takes the average value
                serial1_reading = ser1.read(2)                              # Micrometer sends three bytes, this reads the MSB and LSB
                serial1_STATUS = ser1.read(1)                               # Third byte reads the status
                reading1 = int.from_bytes(serial1_reading, byteorder = 'big', signed=False)     # Converts byte to integer
                reading1_avg += reading1                                    # Stores sum to be used to take average
                n = n + 1                                                   # Iteration
            reading1_avg = reading1_avg / 16                                # Used to average the results
            reading1_avg = (reading1_avg  / 0.4375) / 1000                  # Conversion to mmm    
            reading1_avg = round(reading1_avg, ndigits = 3)                 # sets number of decimal places to 3




            # Serial 2
            ser2.write(b'\x14')                                             # The 4 represents 2^4 captures from the micrometer sensor
            i = 0                                                           # Indexing variable for sensor 2
            reading2_avg = 0                                                # Variable used to obtain our measurement value

            while i < 16:                                                   # 16 readings since 2^4, takes the average value
                serial2_reading = ser2.read(2)                              # Micrometer sends three bytes, this reads the MSB and LSB
                serial2_STATUS = ser2.read(1)                               # Third byte reads the status
                reading2 = int.from_bytes(serial2_reading, byteorder = 'big', signed=False)     # Converts byte to integer
                reading2_avg += reading2                                    # Stores sum to be used to take average
                i = i + 1                                                   # Iteration
            reading2_avg = reading2_avg / 16                                # Used to average the results
            reading2_avg = (reading2_avg  / 0.4375) / 1000                  # Conversion to mm
            reading2_avg = round(reading2_avg, ndigits = 3)                 # Sets number of decimal places
            print(reading1_avg, reading2_avg)
            time.sleep(0.5)                    # OMIT when robot commands are added



            if reading1_avg != 0 and reading2_avg != 0:             # Checks if both sensors have been activated
                if raised_flag1 == 1:                               # Checks if readt flag has been raised
                    state = 3                                      
                elif raised_flag2 == 1: 
                    state = 4                                       # Goes to this state to raise a flag
                else:
                    state = 2


        elif state == 2:
            # State 2: Raise flag
            time.sleep(2)                                           # Halts program for 2 second. Used to wait for the end effector to minimize vibrations 
            raised_flag1 = 1                                        # Raises proceed 
            state = 1                                               # Assigns state = 1 to take another measurement

        elif state == 3: 
            raised_flag1 = 0 
            raised_flag2 = 1
            print('adding to 1')
            reading1_list1.append(reading1_avg)
            reading2_list1.append(reading2_avg)
            print("Pic 1 -   1: ", reading1_avg, "              2: ",reading2_avg)        # Provides update of measurment values onto the output screen
            time.sleep(0.5)                               #################################### CHANGE BACK TO 5
            state = 1


        elif state == 4:     
            # State 4: Add values to list
            raised_flag1 = 0                                        # Resets the flag
            raised_flag2 = 0
            print('adding to 2')
            reading1_list2.append(reading1_avg)                      # Adds measurement values from sensor 1 to the list
            reading2_list2.append(reading2_avg)                      # Adds measurement values from sensor 2 to the list
            print("Pic 2 -   1: ", reading1_avg, "              2: ",reading2_avg)        # Provides update of measurment values onto the output screen
            time.sleep(3)                                     ######CHANGE BACK       # Paues screening for 5 seconds to prevent cluster of zero values on the screen
            print('Cycle(s): ', len(reading1_list2))                 # Used to show what cycle we are on
            if len(reading1_list2) == 500:                          # Desired number of cycles used in data
                workbook = openpyxl.Workbook()                       # Opens desired workbook
                sheet = workbook.active                              # Used to activate excel sheet

                # Add data to the worksheet
                for row_idx, (value1, value2, value3, value4) in enumerate(zip(reading1_list1, reading2_list1, reading1_list2, reading2_list2), start=1):
                    sheet.cell(row=row_idx, column=2, value=value1)                     # Uploads reading 1 values
                    sheet.cell(row=row_idx, column=3, value=value2)                     # Uploads reading 2 values
                    sheet.cell(row=row_idx, column=4, value=value3)
                    sheet.cell(row=row_idx, column=5, value=value4)

                # Specify the Excel file path
                excel_file_path = 'Robot_Laser_Measurement.xlsx'                        # Desired excel file path 

                # Save the workbook to a file
                workbook.save(excel_file_path)                                          # Saves excel file

                print(f'Data exported to {excel_file_path}')                            # Confirmation statement 
                break
            state = 0
        else:
            state = 0


if __name__ == "__main__":
    main()

